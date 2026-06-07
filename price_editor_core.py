#!/usr/bin/env python3
"""
Shopee Mass Price Editor — Core library
Domain logic only. Import from this module in GUI or CLI wrappers.

Public API:
    ProductRow, parse_variation(), calc_price_pct(), calc_price_flat()
    load_file(), save_file(), save_summary_report(), make_output_path()
    build_changes(), merge_changes(), apply_adj()
    unique_products(), price_range(), get_fill()
    THAI_COLORS, SHOPEE_MIN, SHOPEE_MAX, HEADER_ROWS
    SHOEPE_BUILD_LABEL
"""

import sys
import os
import re
import io
import zipfile
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.styles.fills import FILL_SOLID
except ImportError:
    print("Missing dependency: pip install openpyxl")
    sys.exit(1)

# ── Constants ───
BUILD_LABEL = "SHOPEE PRICE EDITOR V1"
SHOPEE_MIN = 1
SHOPEE_MAX = 500_000
HEADER_ROWS = 6
COL_PRODUCT_ID   = 1
COL_PRODUCT_NAME = 2
COL_VARIATION_ID = 3
COL_VARIATION    = 4
COL_PARENT_SKU   = 5
COL_SKU          = 6
COL_PRICE        = 7

# ── Thai color translations ──
THAI_COLORS = {
    "ขาว": "white", "ดำ": "black", "แดง": "red", "น้ำเงิน": "blue",
    "เขียว": "green", "เหลือง": "yellow", "ส้ม": "orange", "ม่วง": "purple",
    "ชมพู": "pink", "เทา": "gray", "น้ำตาล": "brown", "ครีม": "cream",
    "เบจ": "beige", "ฟ้า": "sky blue", "แดงสว่าง": "bright red",
    "แดงมารูน": "maroon", "เขียวละมุน": "soft green", "เขียวเข้ม": "dark green",
    "ฟ้าอ่อน": "light blue", "ชมพูอ่อน": "light pink", "ม่วงอ่อน": "lavender",
    "เทาเข้ม": "dark gray", "กรมท่า": "navy", "ส้มอ่อน": "light orange",
    "มาน่า": "mana", "เงิน": "silver", "ทอง": "gold", "ทองเหลือง": "bronze",
    "กุหลาบ": "rose", "ไวน์": "wine", "มัสตาร์ด": "mustard", "ไพเปอร์": "pepper",
    "สโมกกี้": "smokey", "เทอร์ราคอตต้า": "terracotta", "เขียวมิ้นต์": "mint green",
    "ชา": "tea", "อมชมพู": "dusty pink", "กรีเดน": "olive", "เทาเบา": "light gray",
    "ขาวนม": "milky white", "ดำสนิท": "pure black", "น้ำเงินเข้ม": "dark blue",
    "เขียวมรกต": "emerald", "ฟ้าคม": "royal blue", "ม่วงเข้ม": "dark purple",
    "สตรอว์เบอร์รี่": "strawberry", "บลูเบอร์รี่": "blueberry", "ลาเวนเดอร์": "lavender",
    "แดงอิฐ": "brick red", "ส้มโอ": "tangerine", "เหลืองทอง": "golden yellow",
    "เขียวป่า": "forest green", "น้ำเงินโทน": "tonal blue", "ดำมืด": "charcoal",
    "เทากลาง": "medium gray", "ขาวบริสุทธิ์": "pure white", "ชมพูพาล": "baby pink",
    "สีน้ำเงิน": "navy blue", "สีกากี": "khaki", "สีน้ำตาล": "brown", "สีเทา": "gray",
    "สีชมพู": "pink", "สีเขียว": "green", "สีแดง": "red", "สีเหลือง": "yellow",
}

# ── Fill helpers ──
def make_fill(hex_color):
    return PatternFill(fill_type=FILL_SOLID, fgColor=hex_color.lstrip("#"))

NO_FILL = PatternFill(fill_type=None)

def get_fill(old_price, new_price):
    if new_price == old_price:
        return NO_FILL, False
    pct = abs((new_price - old_price) / old_price * 100)
    increase = new_price > old_price
    if pct < 2:
        hex_c = "#E8F5E9" if increase else "#FFEBEE"
    elif pct < 5:
        hex_c = "#A5D6A7" if increase else "#EF9A9A"
    elif pct < 10:
        hex_c = "#66BB6A" if increase else "#EF5350"
    elif pct < 20:
        hex_c = "#2E7D32" if increase else "#C62828"
    else:
        hex_c = "#1B5E20" if increase else "#7F0000"
    return make_fill(hex_c), True

CAPPED_FILL = make_fill("#FFE0B2")

# ── Variation parsing ──
_CODE_PATTERN     = re.compile(r'^(.*?)\s+(\d+[A-Z]\S*),(.*)$')
_CODE_PATTERN_REV = re.compile(r'^(\d+[A-Z]\S*)\s+(.*?),(.*)$')
_CODE_PATTERN_NUM = re.compile(r'^(.*?)\s+(\d+),(.*)$')
_NO_SKU_PATTERN   = re.compile(r'^([^,]+),(.+)$')
_SIZE_LEADING     = re.compile(r'^(\d+)')
_UNIT_YEAR        = re.compile(r'ปี')
_UNIT_MONTH       = re.compile(r'เดือน')
_PANE_FIX         = re.compile(r'activePane="([^"]+)"')
_SNAKE_TO_CAMEL   = {
    "bottom_left": "bottomLeft", "bottom_right": "bottomRight",
    "top_left": "topLeft", "top_right": "topRight",
}

MAX_XLSX_ENTRIES = 20_000
MAX_XLSX_TOTAL_UNCOMPRESSED = 200 * 1024 * 1024
MAX_XLSX_ENTRY_UNCOMPRESSED = 20 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200

def parse_variation(variation_name):
    """Return (color, size_num, unit) or None."""
    if not variation_name:
        return None
    s = str(variation_name).strip()
    color = size_token = None

    m = _CODE_PATTERN.match(s)
    if m:
        color, size_token = m.group(1).strip(), m.group(3).strip()

    if color is None:
        m = _CODE_PATTERN_REV.match(s)
        if m:
            color, size_token = m.group(2).strip(), m.group(3).strip()

    if color is None:
        m = _CODE_PATTERN_NUM.match(s)
        if m:
            color, size_token = m.group(1).strip(), m.group(3).strip()

    if color is None:
        m = _NO_SKU_PATTERN.match(s)
        if m:
            color, size_token = m.group(1).strip(), m.group(2).strip()

    if color is None or not size_token:
        return None

    sm = _SIZE_LEADING.match(size_token)
    if not sm:
        return None
    size_num = int(sm.group(1))
    is_year  = bool(_UNIT_YEAR.search(size_token)) or size_token.strip().upper().endswith('Y')
    is_month = bool(_UNIT_MONTH.search(size_token))
    if not is_year and not is_month:
        return None
    unit = "ปี" if is_year else "เดือน"
    return color, size_num, unit

# ── Price calculation ──
def calc_price_pct(old_price, pct):
    new_price = round(old_price * (1 + pct / 100), 2)
    capped = False
    if new_price < SHOPEE_MIN:
        new_price = SHOPEE_MIN
        capped = True
    if new_price > SHOPEE_MAX:
        new_price = SHOPEE_MAX
        capped = True
    return new_price, capped

def calc_price_flat(old_price, amount):
    new_price = round(old_price + amount, 2)
    capped = False
    if new_price < SHOPEE_MIN:
        new_price = SHOPEE_MIN
        capped = True
    if new_price > SHOPEE_MAX:
        new_price = SHOPEE_MAX
        capped = True
    return new_price, capped

# ── Excel loading ──
class ProductRow:
    __slots__ = ("row_idx", "product_id", "product_name", "variation_id",
                 "variation_name", "parent_sku", "sku", "price",
                 "color", "size_num", "unit")

    def __init__(self, row_idx, cells):
        self.row_idx       = row_idx
        self.product_id    = cells[0]
        self.product_name  = str(cells[1]) if cells[1] is not None else ""
        self.variation_id  = cells[2]
        self.variation_name= str(cells[3]) if cells[3] is not None else ""
        self.parent_sku    = str(cells[4]) if cells[4] is not None else ""
        self.sku           = str(cells[5]) if cells[5] is not None else ""
        raw_price = cells[6]
        try:
            self.price = float(raw_price)
        except (TypeError, ValueError):
            self.price = None
        parsed = parse_variation(self.variation_name)
        if parsed:
            self.color, self.size_num, self.unit = parsed
        else:
            self.color = self.size_num = self.unit = None

def _repair_xlsx(path):
    def _fix_pane(m):
        val = m.group(1)
        return f'activePane="{_SNAKE_TO_CAMEL.get(val, val)}"'
    buf = io.BytesIO()
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                text = data.decode("utf-8", errors="replace")
                text = _PANE_FIX.sub(_fix_pane, text)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    buf.seek(0)
    return buf

def _validate_xlsx_archive(path):
    total_uncompressed = 0
    with zipfile.ZipFile(path, "r") as zf:
        infos = zf.infolist()
        if len(infos) > MAX_XLSX_ENTRIES:
            raise ValueError("XLSX has too many entries.")
        for info in infos:
            if info.file_size < 0 or info.compress_size < 0:
                raise ValueError("XLSX has invalid entry sizes.")
            if info.file_size > MAX_XLSX_ENTRY_UNCOMPRESSED:
                raise ValueError(f"XLSX entry too large: {info.filename}")
            if info.compress_size > 0:
                ratio = info.file_size / info.compress_size
                if ratio > MAX_XLSX_COMPRESSION_RATIO:
                    raise ValueError(f"XLSX entry compression ratio too high: {info.filename}")
            total_uncompressed += info.file_size
            if total_uncompressed > MAX_XLSX_TOTAL_UNCOMPRESSED:
                raise ValueError("XLSX is too large after decompression.")

def load_file(path):
    _validate_xlsx_archive(path)
    try:
        wb = load_workbook(path)
    except ValueError:
        fixed = _repair_xlsx(path)
        wb = load_workbook(fixed)
    ws = wb.active
    rows = []
    for row_idx, vals in enumerate(
        ws.iter_rows(min_row=HEADER_ROWS + 1, max_col=15, values_only=True),
        start=HEADER_ROWS + 1
    ):
        vals = list(vals)
        if len(vals) < 15:
            vals.extend([None] * (15 - len(vals)))
        pr = ProductRow(row_idx, vals)
        rows.append(pr)
    return wb, ws, rows

def unique_products(rows):
    return len(set(r.product_id for r in rows))

def price_range(rows):
    prices = [r.price for r in rows if r.price is not None]
    return (min(prices), max(prices)) if prices else (None, None)

# ── Adjustment logic ──
def apply_adj(old_price, value, is_pct):
    if is_pct:
        return calc_price_pct(old_price, value)
    else:
        return calc_price_flat(old_price, value)

def build_changes(rows, row_filter, value, is_pct, min_price=None):
    """Return changes dict: row_idx -> (old, new, capped)

    min_price: skip SKUs below this price (for cheap-product filter).
    """
    changes = {}
    for r in rows:
        if r.price is None:
            continue
        if min_price is not None and r.price < min_price:
            continue
        if row_filter(r):
            new_p, capped = apply_adj(r.price, value, is_pct)
            changes[r.row_idx] = (r.price, new_p, capped)
    return changes

def merge_changes(existing, new_changes):
    """Merge new_changes into existing, chaining from latest new price."""
    for idx, (old_p, new_p, capped) in new_changes.items():
        if idx in existing:
            orig_old = existing[idx][0]
            existing[idx] = (orig_old, new_p, capped)
        else:
            existing[idx] = (old_p, new_p, capped)
    return existing

# ── Tag helpers (standardized format: {group}_{sign}{abs_val}{pct|bht}) ──
def make_tag(group, value, is_pct):
    sign = "p" if value >= 0 else "m"
    suffix = "pct" if is_pct else "bht"
    return f"{group}_{sign}{abs(value):.0f}{suffix}"

# ── Save helpers ──
def make_output_path(input_path, suffix="updated", edit_label=""):
    stem = os.path.splitext(os.path.basename(input_path))[0]
    ts   = datetime.now().strftime("%Y-%m-%d_%H-%M")
    if edit_label:
        base = f"{stem}_{BUILD_LABEL}_{edit_label}_{ts}.xlsx"
    else:
        base = f"{stem}_{BUILD_LABEL}_{suffix}_{ts}.xlsx"
    out  = os.path.join(os.path.dirname(os.path.abspath(input_path)), base)
    counter = 2
    while os.path.exists(out):
        if edit_label:
            out = os.path.join(os.path.dirname(os.path.abspath(input_path)),
               f"{stem}_{BUILD_LABEL}_{edit_label}_{ts}_{counter}.xlsx")
        else:
            out = os.path.join(os.path.dirname(os.path.abspath(input_path)),
               f"{stem}_{BUILD_LABEL}_{suffix}_{ts}_{counter}.xlsx")
        counter += 1
    return out

def save_file(wb, ws, rows, changes, output_path):
    bold_font  = Font(bold=True)
    plain_font = Font(bold=False)
    for r in rows:
        cell = ws.cell(row=r.row_idx, column=COL_PRICE)
        if r.row_idx in changes:
            old_p, new_p, capped = changes[r.row_idx]
            if new_p != old_p:
                cell.value = new_p
                if capped:
                    cell.fill = CAPPED_FILL
                else:
                    fill, changed = get_fill(old_p, new_p)
                    cell.fill = fill
                cell.font = bold_font
            else:
                cell.fill = NO_FILL
                cell.font = plain_font
    wb.save(output_path)

def save_summary_report(original_prices, current_prices, rows, output_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Price Change Summary"

    headers = ["Product Name", "Original Price (฿)", "New Price (฿)", "+/- %", "+/- (฿)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = make_fill("#4472C4")

    data_row = 2
    for r in rows:
        if r.row_idx not in original_prices:
            continue
        orig = original_prices[r.row_idx]
        curr = current_prices.get(r.row_idx, orig)
        if orig == curr:
            continue

        ws.cell(row=data_row, column=1, value=r.product_name)
        ws.cell(row=data_row, column=2, value=orig)
        ws.cell(row=data_row, column=3, value=curr)

        diff_pct = ((curr - orig) / orig * 100) if orig else 0
        diff_abs = curr - orig

        ws.cell(row=data_row, column=4, value=diff_pct)
        ws.cell(row=data_row, column=5, value=diff_abs)

        fill, _ = get_fill(orig, curr)
        ws.cell(row=data_row, column=3).fill = fill

        pct_cell = ws.cell(row=data_row, column=4)
        if diff_pct > 0:
            pct_cell.font = Font(color="1B5E20", bold=True)
        elif diff_pct < 0:
            pct_cell.font = Font(color="7F0000", bold=True)

        abs_cell = ws.cell(row=data_row, column=5)
        if diff_abs > 0:
            abs_cell.font = Font(color="1B5E20", bold=True)
        elif diff_abs < 0:
            abs_cell.font = Font(color="7F0000", bold=True)

        data_row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14

    stem = os.path.splitext(os.path.basename(output_path))[0]
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    summary_path = os.path.join(os.path.dirname(output_path), f"{stem}_summary_{ts}.xlsx")
    wb.save(summary_path)
    return summary_path

# ── Pre-built row filter factories (for GUI use) ──
def make_size_filter(selected_sizes):
    """Return row_filter for (size_num, unit) in selected_sizes."""
    ss = set(selected_sizes)
    return lambda r: (r.size_num, r.unit) in ss

def make_color_filter(selected_colors):
    """Return row_filter for r.color in selected_colors."""
    sc = set(selected_colors)
    return lambda r: r.color in sc

def make_sku_filter(sku_token):
    """Return row_filter for r.sku containing sku_token."""
    return lambda r: sku_token in r.sku

# ── Search and price filter helpers ──
def matches_search(row, token):
    """Return True when token matches product name, SKU, or parent SKU."""
    token = str(token or "").strip()
    if not token:
        return True
    token_lower = token.casefold()
    fields = (row.product_name, row.sku, row.parent_sku)
    return any(token_lower in str(field or "").casefold() for field in fields)

def make_search_price_filter(token="", min_price=None, max_price=None, price_source=None):
    """Return row_filter for text search plus optional price bounds."""
    def row_filter(row):
        if not matches_search(row, token):
            return False

        if min_price is None and max_price is None:
            return True

        price = None
        if price_source is not None:
            price = price_source.get(row.row_idx)
        if price is None:
            price = row.price
        if price is None:
            return False
        if min_price is not None and price < min_price:
            return False
        if max_price is not None and price > max_price:
            return False
        return True

    return row_filter

# Tests
def run_tests():
    """Run self-tests. Call with: python price_editor_core.py --test"""
    failures = 0
    passed = 0

    def assert_eq(desc, got, expected):
        nonlocal failures, passed
        ok = abs(got - expected) < 0.001 if isinstance(got, float) else got == expected
        status = "PASS" if ok else "FAIL"
        # Use ascii() to avoid encoding issues on Windows cp1252
        got_s = ascii(got)
        exp_s = ascii(expected)
        print(f"  [{status}] {desc}: got={got_s}  expected={exp_s}")
        if ok:
            passed += 1
        else:
            failures += 1

    # price calc pct
    p, c = calc_price_pct(289, 7)
    assert_eq("pct +7 on 289", p, 309.23)
    p, c = calc_price_pct(289, -5)
    assert_eq("pct -5 on 289", p, 274.55)
    p, c = calc_price_pct(1, -50)
    assert_eq("pct clamped to min", p, 1.0)
    _, capped = calc_price_pct(490000, 10)
    assert_eq("pct clamped to max", capped, True)

    # flat calc
    p, c = calc_price_flat(289, 20)
    assert_eq("flat +20 on 289", p, 309.0)
    p, c = calc_price_flat(289, -15)
    assert_eq("flat -15 on 289", p, 274.0)

    # variation parsing
    r = parse_variation("ขาว 3469S-45,6ปี 20-22kg")
    assert_eq("parse year size", r[1], 6)
    assert_eq("parse year unit", r[2], "ปี")
    r = parse_variation("ดำ 3469S-18,24เดือน 12-13kg")
    assert_eq("parse month size", r[1], 24)
    assert_eq("parse month unit", r[2], "เดือน")

    # build_changes with min_price
    class MR:
        def __init__(self, ri, pr, color=None, size_num=None, unit=None, sku=""):
            self.row_idx, self.price = ri, pr
            self.color, self.size_num, self.unit = color, size_num, unit
            self.sku = sku
            self.product_id, self.product_name = f"P{ri}", f"Prod {ri}"
            self.variation_id, self.variation_name = None, ""
            self.parent_sku = ""

    test_rows = [MR(i, 50 + i * 10) for i in range(1, 6)]  # 60,70,80,90,100
    changes = build_changes(test_rows, lambda r: True, -10, False, min_price=70)
    assert_eq("min_price filter: 4 rows (70,80,90,100)", len(changes), 4)

    # search and price filtering
    thai_name = "\u0e40\u0e2a\u0e37\u0e49\u0e2d\u0e40\u0e14\u0e47\u0e01\u0e2a\u0e35\u0e02\u0e32\u0e27"
    search_rows = [
        MR(10, 120, sku="ABC-001"),
        MR(11, 250, sku="xyz-002"),
        MR(12, None, sku="NO-PRICE"),
    ]
    search_rows[0].product_name = thai_name
    search_rows[0].parent_sku = "PARENT-THAI"
    search_rows[1].product_name = "Dress"
    search_rows[1].parent_sku = "PARENT-XYZ"
    search_rows[2].product_name = "No price"
    search_rows[2].parent_sku = "MISSING"

    assert_eq("Thai product search", matches_search(search_rows[0], "\u0e40\u0e2a\u0e37\u0e49\u0e2d"), True)
    assert_eq("SKU search case-insensitive", matches_search(search_rows[1], "XYZ"), True)
    assert_eq("parent SKU search", matches_search(search_rows[0], "parent-thai"), True)

    f = make_search_price_filter(min_price=100)
    assert_eq("price filter min only", len([r for r in search_rows if f(r)]), 2)
    f = make_search_price_filter(max_price=150)
    assert_eq("price filter max only", len([r for r in search_rows if f(r)]), 1)
    f = make_search_price_filter(min_price=100, max_price=200)
    assert_eq("price filter min and max", len([r for r in search_rows if f(r)]), 1)

    original_prices = {10: 120, 11: 250, 12: 175}
    search_rows[1].price = 150
    f = make_search_price_filter(min_price=200, price_source=original_prices)
    assert_eq("price filter uses provided source", len([r for r in search_rows if f(r)]), 1)

    # tag format
    tag = make_tag("size", -7, True)
    assert_eq("tag size -7pct", tag, "size_m7pct")
    tag = make_tag("color", 15, False)
    assert_eq("tag color +15bht", tag, "color_p15bht")

    print(f"\n  {'='*46}")
    if failures == 0:
        print(f"  All tests passed. ({passed} assertions)")
    else:
        print(f"  {failures} test(s) failed.")
    print(f"{'='*50}\n")
    return failures == 0

if __name__ == "__main__":
    if "--test" in sys.argv:
        ok = run_tests()
        sys.exit(0 if ok else 1)
    else:
        print("price_editor_core — domain library. Import from GUI.")
        print("Usage: python price_editor_core.py --test")
