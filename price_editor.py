#!/usr/bin/env python3
"""
Shopee Mass Price Editor
Bulk-adjust prices in Shopee Thailand mass-update Excel files.
Usage: price_editor.exe [--test]
"""

# ── Version / build label ──────────────────────────────
BUILD_LABEL = "SHOPEE PRICE EDITOR V1"

import sys
import os
import re
import io
import zipfile

# Force UTF-8 output on Windows so Thai text and box-drawing chars render correctly
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf-8-sig"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
from datetime import datetime

try:
    import msvcrt
    _HAS_MSVCRT = True
except ImportError:
    msvcrt = None
    _HAS_MSVCRT = False

def _getch():
    """Get a single character without echo."""
    if _HAS_MSVCRT:
        return msvcrt.getch()
    return None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.styles.fills import FILL_SOLID
except ImportError:
    print("Missing dependency: pip install openpyxl")
    sys.exit(1)

# ── ANSI helpers ──────────────────────────────────────────────────────────────
GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
CYAN   = "\033[96m"
BOLD   = "\033[1m"
RESET  = "\033[0m"
DIM    = "\033[2m"

# ── Constants ─────────────────────────────────────────────────────────────────
SHOPEE_MIN = 1
SHOPEE_MAX = 500_000
HEADER_ROWS = 6
COL_PRODUCT_ID   = 1   # A
COL_PRODUCT_NAME = 2   # B
COL_VARIATION_ID = 3   # C
COL_VARIATION    = 4   # D
COL_PARENT_SKU   = 5   # E
COL_SKU          = 6   # F
COL_PRICE        = 7   # G

# ── Thai color translations ───────────────────────────────────────────────────
THAI_COLORS = {
    "ขาว": "white",
    "ดำ": "black",
    "แดง": "red",
    "น้ำเงิน": "blue",
    "เขียว": "green",
    "เหลือง": "yellow",
    "ส้ม": "orange",
    "ม่วง": "purple",
    "ชมพู": "pink",
    "เทา": "gray",
    "น้ำตาล": "brown",
    "ครีม": "cream",
    "เบจ": "beige",
    "ฟ้า": "sky blue",
    "แดงสว่าง": "bright red",
    "แดงมารูน": "maroon",
    "เขียวละมุน": "soft green",
    "เขียวเข้ม": "dark green",
    "ฟ้าอ่อน": "light blue",
    "ชมพูอ่อน": "light pink",
    "ม่วงอ่อน": "lavender",
    "เทาเข้ม": "dark gray",
    "กรมท่า": "navy",
    "ส้มอ่อน": "light orange",
}

# ── Fill helpers ──────────────────────────────────────────────────────────────
def make_fill(hex_color):
    return PatternFill(fill_type=FILL_SOLID, fgColor=hex_color.lstrip("#"))

NO_FILL = PatternFill(fill_type=None)

def get_fill(old_price, new_price):
    if new_price == old_price:
        return NO_FILL, False
    pct = abs((new_price - old_price) / old_price * 100)
    increase = new_price > old_price
    if pct < 2:
        hex_c = "#E8F5E9" if increase else "#FFEBEE"
    elif pct < 5:
        hex_c = "#A5D6A7" if increase else "#EF9A9A"
    elif pct < 10:
        hex_c = "#66BB6A" if increase else "#EF5350"
    elif pct < 20:
        hex_c = "#2E7D32" if increase else "#C62828"
    else:
        hex_c = "#1B5E20" if increase else "#7F0000"
    return make_fill(hex_c), True

CAPPED_FILL = make_fill("#FFE0B2")

# ── Variation parsing ─────────────────────────────────────────────────────────
_CODE_PATTERN     = re.compile(r'^(.*?)\s+(\d+[A-Z]\S*),(.*)$')   # Format 2: COLOR SKU,SIZE WEIGHT
_CODE_PATTERN_REV = re.compile(r'^(\d+[A-Z]\S*)\s+(.*?),(.*)$')   # Format 1: SKU COLOR,SIZE WEIGHT
_CODE_PATTERN_NUM = re.compile(r'^(.*?)\s+(\d+),(.*)$')            # Format 4: COLOR NUM,SIZE (pure-number code)
_NO_SKU_PATTERN   = re.compile(r'^([^,]+),(.+)$')                  # Format 3: LABEL,SIZE WEIGHT (no code)
_SIZE_LEADING     = re.compile(r'^(\d+)')
_UNIT_YEAR        = re.compile(r'ปี')
_UNIT_MONTH       = re.compile(r'เดือน')
_PANE_FIX         = re.compile(r'activePane="([^"]+)"')
_SNAKE_TO_CAMEL   = {
    "bottom_left":  "bottomLeft",
    "bottom_right": "bottomRight",
    "top_left":     "topLeft",
    "top_right":    "topRight",
}

# Guardrails against zip bombs / malformed archives.
MAX_XLSX_ENTRIES = 20_000
MAX_XLSX_TOTAL_UNCOMPRESSED = 200 * 1024 * 1024   # 200 MiB
MAX_XLSX_ENTRY_UNCOMPRESSED = 20 * 1024 * 1024    # 20 MiB
MAX_XLSX_COMPRESSION_RATIO = 200

def parse_variation(variation_name):
    """Return (color, size_num, unit) or None on failure."""
    if not variation_name:
        return None
    s = str(variation_name).strip()
    color = size_token = None

    m = _CODE_PATTERN.match(s)                          # Format 2 (most common)
    if m:
        color, size_token = m.group(1).strip(), m.group(3).strip()

    if color is None:
        m = _CODE_PATTERN_REV.match(s)                  # Format 1: SKU first
        if m:
            color, size_token = m.group(2).strip(), m.group(3).strip()

    if color is None:
        m = _CODE_PATTERN_NUM.match(s)                  # Format 4: pure-number code
        if m:
            color, size_token = m.group(1).strip(), m.group(3).strip()

    if color is None:
        m = _NO_SKU_PATTERN.match(s)                    # Format 3: no code at all
        if m:
            color, size_token = m.group(1).strip(), m.group(2).strip()

    if color is None or not size_token:
        return None

    sm = _SIZE_LEADING.match(size_token)
    if not sm:
        return None
    size_num = int(sm.group(1))
    is_year  = bool(_UNIT_YEAR.search(size_token)) or size_token.strip().upper().endswith('Y')
    is_month = bool(_UNIT_MONTH.search(size_token))
    if not is_year and not is_month:
        return None   # physical size / dimension, not an age — ignore
    unit = "ปี" if is_year else "เดือน"
    return color, size_num, unit

# ── Price calculation ─────────────────────────────────────────────────────────
def calc_price_pct(old_price, pct):
    new_price = round(old_price * (1 + pct / 100), 2)
    capped = False
    if new_price < SHOPEE_MIN:
        new_price = SHOPEE_MIN
        capped = True
    if new_price > SHOPEE_MAX:
        new_price = SHOPEE_MAX
        capped = True
    return new_price, capped

def calc_price_flat(old_price, amount):
    new_price = round(old_price + amount, 2)
    capped = False
    if new_price < SHOPEE_MIN:
        new_price = SHOPEE_MIN
        capped = True
    if new_price > SHOPEE_MAX:
        new_price = SHOPEE_MAX
        capped = True
    return new_price, capped

# ── Excel loading ─────────────────────────────────────────────────────────────
class ProductRow:
    __slots__ = ("row_idx", "product_id", "product_name", "variation_id",
                 "variation_name", "parent_sku", "sku", "price",
                 "color", "size_num", "unit")

    def __init__(self, row_idx, cells):
        self.row_idx       = row_idx
        self.product_id    = cells[0]
        self.product_name  = str(cells[1]) if cells[1] is not None else ""
        self.variation_id  = cells[2]
        self.variation_name= str(cells[3]) if cells[3] is not None else ""
        self.parent_sku    = str(cells[4]) if cells[4] is not None else ""
        self.sku           = str(cells[5]) if cells[5] is not None else ""
        raw_price = cells[6]
        try:
            self.price = float(raw_price)
        except (TypeError, ValueError):
            self.price = None
        parsed = parse_variation(self.variation_name)
        if parsed:
            self.color, self.size_num, self.unit = parsed
        else:
            self.color = self.size_num = self.unit = None

def _repair_xlsx(path):
    """
    Shopee's mass-update export writes activePane with snake_case values
    (e.g. "bottom_left") which violates the OOXML spec — openpyxl rejects them.
    We patch every worksheet XML in-memory and return a BytesIO of the fixed zip.
    """
    def _fix_pane(m):
        val = m.group(1)
        return f'activePane="{_SNAKE_TO_CAMEL.get(val, val)}"'

    buf = io.BytesIO()
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                text = data.decode("utf-8", errors="replace")
                text = _PANE_FIX.sub(_fix_pane, text)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    buf.seek(0)
    return buf

def _validate_xlsx_archive(path):
    total_uncompressed = 0
    with zipfile.ZipFile(path, "r") as zf:
        infos = zf.infolist()
        if len(infos) > MAX_XLSX_ENTRIES:
            raise ValueError("XLSX has too many entries.")
        for info in infos:
            if info.file_size < 0 or info.compress_size < 0:
                raise ValueError("XLSX has invalid entry sizes.")
            if info.file_size > MAX_XLSX_ENTRY_UNCOMPRESSED:
                raise ValueError(f"XLSX entry too large: {info.filename}")
            if info.compress_size > 0:
                ratio = info.file_size / info.compress_size
                if ratio > MAX_XLSX_COMPRESSION_RATIO:
                    raise ValueError(f"XLSX entry compression ratio too high: {info.filename}")
            total_uncompressed += info.file_size
            if total_uncompressed > MAX_XLSX_TOTAL_UNCOMPRESSED:
                raise ValueError("XLSX is too large after decompression.")


def load_file(path):
    _validate_xlsx_archive(path)
    try:
        wb = load_workbook(path)
    except ValueError:
        fixed = _repair_xlsx(path)
        wb = load_workbook(fixed)
    ws = wb.active
    rows = []
    for row_idx, vals in enumerate(
        ws.iter_rows(min_row=HEADER_ROWS + 1, max_col=15, values_only=True),
        start=HEADER_ROWS + 1
    ):
        vals = list(vals)
        if len(vals) < 15:
            vals.extend([None] * (15 - len(vals)))
        pr = ProductRow(row_idx, vals)
        rows.append(pr)
    return wb, ws, rows

def unique_products(rows):
    seen = set()
    for r in rows:
        seen.add(r.product_id)
    return len(seen)

def price_range(rows):
    prices = [r.price for r in rows if r.price is not None]
    if not prices:
        return None, None
    return min(prices), max(prices)

# ── Input parsing helpers ─────────────────────────────────────────────────────
def parse_adj(raw):
    """Parse adjustment string. Returns (float_value, is_pct) or raises ValueError."""
    raw = raw.strip()
    is_pct = raw.endswith('%')
    num_str = raw.rstrip('%').strip()
    val = float(num_str)
    return val, is_pct

# ── Output filename ───────────────────────────────────────────────────────────
def make_output_path(input_path, suffix="updated", edit_label=""):
    stem = os.path.splitext(os.path.basename(input_path))[0]
    ts   = datetime.now().strftime("%Y-%m-%d_%H-%M")
    if edit_label:
        base = f"{stem}_{BUILD_LABEL}_{edit_label}_{ts}.xlsx"
    else:
        base = f"{stem}_{BUILD_LABEL}_{suffix}_{ts}.xlsx"
    out  = os.path.join(os.path.dirname(os.path.abspath(input_path)), base)
    counter = 2
    while os.path.exists(out):
        out = os.path.join(os.path.dirname(os.path.abspath(input_path)),
                           f"{stem}_{BUILD_LABEL}_{edit_label}_{ts}_{counter}.xlsx" if edit_label else
                           f"{stem}_{BUILD_LABEL}_{suffix}_{ts}_{counter}.xlsx")
        counter += 1
    return out

# ── Preview ───────────────────────────────────────────────────────────────────
def show_preview(rows, changes):
    """
    changes: dict[row_idx] = (old_price, new_price, capped)
    """
    changed = [(r, changes[r.row_idx]) for r in rows if r.row_idx in changes
               and changes[r.row_idx][0] != changes[r.row_idx][1]]
    capped_count  = sum(1 for _, (o, n, cap) in changed if cap)
    skipped_count = sum(1 for r in rows if r.price is None)

    print(f"\n{CYAN}{'─'*60}{RESET}")
    print(f"  {BOLD}PREVIEW — {len(changed)} SKUs affected{RESET}")
    print(f"{CYAN}{'─'*60}{RESET}")

    header = f"  {'Product':<20}  {'Variation':<30}  {'Old ฿':>7}  {'New ฿':>7}  {'Change':<16}"
    print(header)
    print(f"  {'─'*20}  {'─'*30}  {'─'*7}  {'─'*7}  {'─'*16}")

    display = changed[:20]
    for r, (old_p, new_p, cap) in display:
        pname = (r.product_name[:17] + "...") if len(r.product_name) > 20 else r.product_name
        vname = (r.variation_name[:27] + "...") if len(r.variation_name) > 30 else r.variation_name
        diff  = new_p - old_p
        pct   = diff / old_p * 100 if old_p else 0
        sign  = "+" if diff >= 0 else ""
        color = GREEN if diff > 0 else RED
        cap_mark = " [CAPPED]" if cap else ""
        change_str = f"{sign}{diff:.0f} ({sign}{pct:.1f}%){cap_mark}"
        print(f"  {pname:<20}  {vname:<30}  {old_p:>7.0f}  {new_p:>7.0f}  {color}{change_str}{RESET}")

    if len(changed) > 20:
        print(f"\n  {DIM}[showing 20 of {len(changed)} — all changes follow the same pattern]{RESET}")

    # Warnings — 5× ratio
    product_prices = {}
    for r in rows:
        p = changes.get(r.row_idx, (r.price, r.price, False))[1] if r.row_idx in changes else r.price
        if p is None:
            continue
        pid = r.product_id
        if pid not in product_prices:
            product_prices[pid] = {"name": r.product_name, "prices": []}
        product_prices[pid]["prices"].append(p)

    warnings = []
    for pid, info in product_prices.items():
        ps = info["prices"]
        if len(ps) < 2:
            continue
        ratio = max(ps) / min(ps) if min(ps) > 0 else 0
        if ratio > 5:
            warnings.append((info["name"], min(ps), max(ps), ratio))

    if warnings:
        print()
        for name, mn, mx, ratio in warnings:
            print(f"  {YELLOW}⚠  Warning: Product \"{name[:40]}\" will have prices ranging")
            print(f"     from ฿{mn:.0f} to ฿{mx:.0f} after adjustment (ratio {ratio:.1f}×).")
            print(f"     Shopee requires max/min ≤ 5×. This may cause a rejection.{RESET}")

    # Summary
    if changed:
        diffs      = [n - o for _, (o, n, _) in changed]
        avg_diff   = sum(diffs) / len(diffs)
        avg_pct    = sum((n - o) / o * 100 for _, (o, n, _) in changed if o) / len(changed)
        new_prices = [n for _, (_, n, _) in changed]
        all_prices = [changes.get(r.row_idx, (r.price, r.price, False))[1]
                      if r.row_idx in changes else r.price
                      for r in rows if r.price is not None]
        sign = "+" if avg_diff >= 0 else ""

    old_min, old_max = price_range(rows)
    new_min = min(all_prices) if changed else old_min
    new_max = max(all_prices) if changed else old_max

    print(f"\n{CYAN}{'─'*60}{RESET}")
    print(f"  {BOLD}Summary:{RESET}")
    print(f"    SKUs changed:    {len(changed)}")
    if changed:
        color = GREEN if avg_diff >= 0 else RED
        print(f"    Avg change:      {color}{sign}฿{abs(avg_diff):.2f} ({sign}{avg_pct:.1f}%){RESET}")
    print(f"    Price range:     ฿{new_min:.0f} – ฿{new_max:.0f}  (was ฿{old_min:.0f} – ฿{old_max:.0f})")
    print(f"    Capped by limit: {capped_count} SKUs  (Shopee max ฿{SHOPEE_MAX:,})")
    if skipped_count:
        print(f"    Skipped rows:    {skipped_count} (empty/non-numeric price)")
    print(f"{CYAN}{'─'*60}{RESET}")

# ── Save ──────────────────────────────────────────────────────────────────────
def save_file(wb, ws, rows, changes, output_path):
    bold_font  = Font(bold=True)
    plain_font = Font(bold=False)
    for r in rows:
        cell = ws.cell(row=r.row_idx, column=COL_PRICE)
        if r.row_idx in changes:
            old_p, new_p, capped = changes[r.row_idx]
            if new_p != old_p:
                cell.value = new_p
                if capped:
                    cell.fill = CAPPED_FILL
                else:
                    fill, changed = get_fill(old_p, new_p)
                    cell.fill = fill
                cell.font = bold_font
            else:
                cell.fill = NO_FILL
                cell.font = plain_font
    wb.save(output_path)

# ── Summary report ────────────────────────────────────────────────────────────
def save_summary_report(original_prices, current_prices, rows, output_path):
    """
    Write single summary XLSX with 5 columns:
    1. Product name
    2. Original price
    3. New price (with color fill)
    4. +/- %
    5. +/- price (THB)
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Price Change Summary"

    # Header row
    headers = ["Product Name", "Original Price (฿)", "New Price (฿)", "+/- %", "+/- (฿)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = make_fill("#4472C4")
        cell.font = Font(bold=True, color="FFFFFF")

    # Data rows
    data_row = 2
    for r in rows:
        if r.row_idx not in original_prices:
            continue
        orig = original_prices[r.row_idx]
        curr = current_prices.get(r.row_idx, orig)
        if orig == curr:
            continue  # skip unchanged

        ws.cell(row=data_row, column=1, value=r.product_name)
        ws.cell(row=data_row, column=2, value=orig)
        ws.cell(row=data_row, column=3, value=curr)

        diff_pct = ((curr - orig) / orig * 100) if orig else 0
        diff_abs = curr - orig

        ws.cell(row=data_row, column=4, value=diff_pct)
        ws.cell(row=data_row, column=5, value=diff_abs)

        # Color code column 3 (new price) based on change
        fill, _ = get_fill(orig, curr)
        ws.cell(row=data_row, column=3).fill = fill

        # Color code column 4 (% change)
        pct_cell = ws.cell(row=data_row, column=4)
        if diff_pct > 0:
            pct_cell.font = Font(color="1B5E20", bold=True)  # dark green
        elif diff_pct < 0:
            pct_cell.font = Font(color="7F0000", bold=True)  # dark red

        # Color code column 5 (abs change)
        abs_cell = ws.cell(row=data_row, column=5)
        if diff_abs > 0:
            abs_cell.font = Font(color="1B5E20", bold=True)
        elif diff_abs < 0:
            abs_cell.font = Font(color="7F0000", bold=True)

        data_row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14

    # Save
    stem = os.path.splitext(os.path.basename(output_path))[0]
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    summary_path = os.path.join(os.path.dirname(output_path), f"{stem}_summary_{ts}.xlsx")

    wb.save(summary_path)
    return summary_path

# ── Interactive helpers ───────────────────────────────────────────────────────
def clear_lines(n):
    """Move cursor up n lines and clear them."""
    for _ in range(n):
        sys.stdout.write("\033[A\033[K")
    sys.stdout.flush()

def select_menu_arrow(options, title):
    """
    Arrow-key menu selection. Returns selected option key or None if cancelled.
    options: list of (key, label, description) tuples
    """
    if not options:
        return None
    selected = 0
    num_rows = len(options) + 2  # title + footer

    def draw():
        # Move up to title line and clear
        sys.stdout.write(f"\033[{num_rows}A\033[J")
        sys.stdout.write(f"  {BOLD}{title}{RESET}\n")
        for i, (key, label, desc) in enumerate(options):
            prefix = f"{GREEN}▶{RESET} " if i == selected else "  "
            sys.stdout.write(f"  {prefix}{BOLD}{key}{RESET}  {label}")
            if desc:
                sys.stdout.write(f"  {DIM}{desc}{RESET}")
            sys.stdout.write("\n")
        sys.stdout.write(f"  {DIM}↑↓ navigate · Enter select · Esc cancel{RESET}\n")
        sys.stdout.flush()

    # Initial draw - don't move up
    sys.stdout.write(f"  {BOLD}{title}{RESET}\n")
    for i, (key, label, desc) in enumerate(options):
        prefix = f"{GREEN}▶{RESET} " if i == selected else "  "
        sys.stdout.write(f"  {prefix}{BOLD}{key}{RESET}  {label}")
        if desc:
            sys.stdout.write(f"  {DIM}{desc}{RESET}")
        sys.stdout.write("\n")
    sys.stdout.write(f"  {DIM}↑↓ navigate · Enter select · Esc cancel{RESET}\n")
    sys.stdout.flush()

    while True:
        if not _HAS_MSVCRT:
            sys.stdout.write("\n  Enter choice: ")
            sys.stdout.flush()
            try:
                return input().strip().lower()
            except (EOFError, KeyboardInterrupt):
                return None

        if msvcrt.kbhit():
            key = msvcrt.getch()
            if key == b'\xe0':
                key = msvcrt.getch()
                if key == b'H':
                    selected = (selected - 1) % len(options)
                    draw()
                elif key == b'P':
                    selected = (selected + 1) % len(options)
                    draw()
            elif key == b'\r':
                return options[selected][0]
            elif key == b'\x1b':
                return None
            elif key in (b'1', b'2', b'3', b'4'):
                for opt_key, label, _ in options:
                    if opt_key.startswith(key.decode()):
                        return opt_key
    return None

def select_sizes_interactive(size_map, sorted_sizes):
    """
    Arrow-key size selection with multi-select via Space.
    Returns list of (size_num, unit) tuples or None if cancelled.
    """
    if not sorted_sizes:
        return None

    selected_idx = 0
    selected_set = set()
    num_rows = len(sorted_sizes) + 3  # title + selected line + footer

    def draw():
        # Move up and clear
        sys.stdout.write(f"\033[{num_rows}A\033[J")
        sys.stdout.write(f"  {BOLD}Select sizes (Space to toggle, Enter confirm):{RESET}\n")
        for i, size in enumerate(sorted_sizes):
            num, unit = size
            label = f"{num} {'months' if unit == 'เดือน' else 'years'}"
            cnt = size_map[size]
            marker = f"{GREEN}●{RESET}" if size in selected_set else f"{DIM}○{RESET}"
            prefix = f"{GREEN}▶{RESET}" if i == selected_idx else " "
            sys.stdout.write(f"  {prefix} {marker} {label:<12} — {cnt} SKUs\n")
        sel_list = ", ".join(f"{s[0]}{'y' if s[1]=='ปี' else 'm'}" for s in sorted(selected_set, key=lambda x: (x[1], x[0]))) if selected_set else "none"
        sys.stdout.write(f"  {DIM}Selected: {sel_list} · ↑↓ navigate · Space toggle · Enter confirm · Esc cancel{RESET}\n")
        sys.stdout.flush()

    # Initial draw
    sys.stdout.write(f"  {BOLD}Select sizes (Space to toggle, Enter confirm):{RESET}\n")
    for i, size in enumerate(sorted_sizes):
        num, unit = size
        label = f"{num} {'months' if unit == 'เดือน' else 'years'}"
        cnt = size_map[size]
        marker = f"{GREEN}●{RESET}" if size in selected_set else f"{DIM}○{RESET}"
        prefix = f"{GREEN}▶{RESET}" if i == selected_idx else " "
        sys.stdout.write(f"  {prefix} {marker} {label:<12} — {cnt} SKUs\n")
    sel_list = ", ".join(f"{s[0]}{'y' if s[1]=='ปี' else 'm'}" for s in sorted(selected_set, key=lambda x: (x[1], x[0]))) if selected_set else "none"
    sys.stdout.write(f"  {DIM}Selected: {sel_list} · ↑↓ navigate · Space toggle · Enter confirm · Esc cancel{RESET}\n")
    sys.stdout.flush()

    while True:
        if not _HAS_MSVCRT:
            sys.stdout.write("\n  Enter sizes (e.g., 3,4,5 or 3y-14y): ")
            sys.stdout.flush()
            try:
                line = input().strip()
                return parse_size_input(line, size_map, sorted_sizes)
            except (EOFError, KeyboardInterrupt):
                return None

        if msvcrt.kbhit():
            key = msvcrt.getch()
            if key == b'\xe0':
                key = msvcrt.getch()
                if key == b'H':
                    selected_idx = (selected_idx - 1) % len(sorted_sizes)
                    draw()
                elif key == b'P':
                    selected_idx = (selected_idx + 1) % len(sorted_sizes)
                    draw()
            elif key == b' ':
                if sorted_sizes[selected_idx] in selected_set:
                    selected_set.remove(sorted_sizes[selected_idx])
                else:
                    selected_set.add(sorted_sizes[selected_idx])
                draw()
            elif key == b'\r':
                if not selected_set:
                    return [sorted_sizes[selected_idx]]
                return list(selected_set)
            elif key == b'\x1b':
                return None
            elif key == b'0' or (key >= b'1' and key <= b'9'):
                digits = [key]
                while msvcrt.kbhit():
                    k = msvcrt.getch()
                    if k in (b'0', b'1', b'2', b'3', b'4', b'5', b'6', b'7', b'8', b'9', b',', b'-', b'y', b'Y', b'm', b'M'):
                        digits.append(k)
                    else:
                        break
                line = b''.join(digits).decode('utf-8', errors='ignore')
                return parse_size_input(line, size_map, sorted_sizes)
    return None

def parse_size_input(line, size_map, sorted_sizes):
    """Parse typed size input like '3,4,5' or '3y-14y' or '3yr-14'."""
    if not line.strip():
        return None

    # Range format: 3y-14y or 3yr-14 or 3-14y
    range_match = re.match(r'^(\d+)\s*(y|yr|years|m|mo|months)?\s*-\s*(\d+)\s*(y|yr|years|m|mo|months)?$', line.strip(), re.IGNORECASE)
    if range_match:
        start_num = int(range_match.group(1))
        start_unit = range_match.group(2)
        end_num = int(range_match.group(3))
        end_unit = range_match.group(4)

        def resolve_unit(raw, fallback=None):
            if not raw:
                return fallback
            raw = raw.lower().strip()
            if raw in ('y', 'yr', 'years'):
                return 'ปี'
            if raw in ('m', 'mo', 'months'):
                return 'เดือน'
            return None

        unit = resolve_unit(start_unit) or resolve_unit(end_unit)
        if not unit:
            # Default to years if numbers are small (1-14)
            unit = 'ปี' if end_num <= 14 else 'เดือน'

        result = [(num, u) for num, u in sorted_sizes if unit == u and start_num <= num <= end_num]
        return result if result else None

    # Comma-separated: 3,4,5 or 3y,4y,5y
    unit = None
    if 'y' in line.lower() and 'm' not in line.lower():
        unit = 'ปี'
    elif 'm' in line.lower():
        unit = 'เดือน'

    nums = re.findall(r'\d+', line)
    if nums:
        result = [(int(n), unit) for n in nums if unit]
        # If no unit specified, try to match existing sizes
        if not unit:
            for n in nums:
                for size in sorted_sizes:
                    if size[0] == int(n):
                        result.append(size)
                        break
        return result if result else None

    return None

def type_size_range(sorted_sizes):
    """Let user type size range like: 3yr-14 -7"""
    sys.stdout.write(f"\n  {BOLD}Enter size range and adjustment:{RESET}\n")
    sys.stdout.write(f"  {DIM}Format: [start][unit]-[end] [adjustment]{RESET}\n")
    sys.stdout.write(f"  {DIM}Examples: 3yr-14 +7%  or  3-14y -5  or  3months-24months +10%{RESET}\n")
    sys.stdout.write("  > ")
    sys.stdout.flush()

    try:
        line = input().strip()
        if not line:
            return None, None

        # Split range and adjustment
        parts = line.rsplit(None, 1)
        if len(parts) < 2:
            print(f"  {RED}Need: size-range adjustment{RESET}")
            return None, None

        range_part = parts[0].strip()
        adj_part = parts[1].strip()

        # Parse range
        range_match = re.match(r'^(\d+)\s*(y|yr|years|m|mo|months)?\s*-\s*(\d+)\s*(y|yr|years|m|mo|months)?$', range_part, re.IGNORECASE)
        if not range_match:
            print(f"  {RED}Invalid range format. Use: 3yr-14 or 3-14y{RESET}")
            return None, None

        start_num = int(range_match.group(1))
        start_unit = range_match.group(2)
        end_num = int(range_match.group(3))
        end_unit = range_match.group(4)

        def resolve_unit(raw, fallback=None):
            if not raw:
                return fallback
            raw = raw.lower().strip()
            if raw in ('y', 'yr', 'years'):
                return 'ปี'
            if raw in ('m', 'mo', 'months'):
                return 'เดือน'
            return None

        unit = resolve_unit(start_unit) or resolve_unit(end_unit)
        if not unit:
            unit = 'ปี' if end_num <= 14 else 'เดือน'

        # Get matching sizes
        selected = [(num, u) for num, u in sorted_sizes if u == unit and start_num <= num <= end_num]
        if not selected:
            print(f"  {YELLOW}No sizes match that range.{RESET}")
            return None, None

        return selected, adj_part
    except (ValueError, OverflowError) as e:
        print(f"  {RED}Error parsing: {e}{RESET}")
        return None, None

# ── Mode helpers ──────────────────────────────────────────────────────────────
def apply_adj(old_price, value, is_pct):
    if is_pct:
        return calc_price_pct(old_price, value)
    else:
        return calc_price_flat(old_price, value)

def build_changes(rows, row_filter, value, is_pct):
    """Return changes dict: row_idx -> (old, new, capped)"""
    changes = {}
    for r in rows:
        if r.price is None:
            continue
        if row_filter(r):
            new_p, capped = apply_adj(r.price, value, is_pct)
            changes[r.row_idx] = (r.price, new_p, capped)
    return changes

def merge_changes(existing, new_changes):
    """Merge new_changes into existing, chaining from latest new price."""
    for idx, (old_p, new_p, capped) in new_changes.items():
        if idx in existing:
            orig_old = existing[idx][0]
            existing[idx] = (orig_old, new_p, capped)
        else:
            existing[idx] = (old_p, new_p, capped)
    return existing

# ── Unit normalisation ────────────────────────────────────────────────────────
_UNIT_ALIASES = {
    "years": "ปี", "year": "ปี", "yrs": "ปี", "yr": "ปี", "y": "ปี", "ปี": "ปี",
    "months": "เดือน", "month": "เดือน", "mos": "เดือน", "mo": "เดือน", "m": "เดือน", "เดือน": "เดือน",
}

def _parse_unit(token):
    return _UNIT_ALIASES.get(token.lower())

# ── Size display ──────────────────────────────────────────────────────────────
def show_sizes(rows):
    size_map = {}
    for r in rows:
        if r.size_num is None:
            continue
        key = (r.size_num, r.unit)
        size_map[key] = size_map.get(key, 0) + 1

    def sort_key(k):
        num, unit = k
        return (0 if unit == "เดือน" else 1, num)

    sorted_sizes = sorted(size_map.keys(), key=sort_key)
    print(f"\n  {BOLD}Available sizes:{RESET}")
    for (num, unit) in sorted_sizes:
        cnt = size_map[(num, unit)]
        label = f"{num} months" if unit == "เดือน" else f"{num} years"
        print(f"    {label:<14} —  {cnt} SKUs")
    return size_map, sorted_sizes

def show_colors(rows):
    color_map = {}
    for r in rows:
        if r.color is None:
            continue
        color_map[r.color] = color_map.get(r.color, 0) + 1

    sorted_colors = sorted(color_map.keys(), key=lambda c: -color_map[c])
    print(f"\n  {BOLD}Available colors:{RESET}")
    for i, c in enumerate(sorted_colors, 1):
        eng = THAI_COLORS.get(c, "")
        eng_str = f"({eng})" if eng else ""
        cnt = color_map[c]
        print(f"    {i}. {c:<25} {eng_str:<16} —  {cnt} SKUs")
    return color_map, sorted_colors

# ── Mode implementations ──────────────────────────────────────────────────────
def mode_all(rows, is_pct, return_val=False):
    kind = "%" if is_pct else "฿"
    val_raw = input(f"  Enter adjustment ({kind}): ").strip()
    try:
        value, detected_pct = parse_adj(val_raw)
        if is_pct:
            detected_pct = True  # override
    except ValueError:
        print(f"  {RED}Invalid input.{RESET}")
        return ({}, "") if return_val else {}
    if is_pct:
        detected_pct = True
    changes = build_changes(rows, lambda r: True, value, detected_pct)
    if return_val:
        return changes, value
    return changes


def mode_by_size(rows, is_pct, return_tag=False):
    size_map, sorted_sizes = show_sizes(rows)
    kind = "%" if is_pct else "฿"

    # Ask user how they want to select sizes
    print(f"\n  {BOLD}How to select sizes?{RESET}")
    print(f"    1. Arrow keys (↑↓) + Space for multi-select")
    print(f"    2. Type range (e.g., 3yr-14)")
    sys.stdout.write("  > ")
    sys.stdout.flush()

    choice = _getch()
    if choice:
        choice = choice.decode('utf-8', errors='ignore')
        print(choice + "\n")
    else:
        choice = input().strip()

    all_changes = {}
    all_values = []
    selected_sizes = []
    value = 0.0

    if choice == '2':
        # Typed range mode
        selected, adj_str = type_size_range(sorted_sizes)
        if not selected or not adj_str:
            return (all_changes, "") if return_tag else all_changes

        try:
            value, detected_pct = parse_adj(adj_str)
            if is_pct:
                detected_pct = True
        except ValueError:
            print(f"  {RED}Invalid adjustment.{RESET}")
            return (all_changes, "") if return_tag else all_changes

        selected_sizes = selected
        selected_set = set(selected_sizes)
        all_changes = build_changes(
            rows,
            lambda r, ss=selected_set: (r.size_num, r.unit) in ss,
            value,
            detected_pct
        )
        all_values.append(value)
        cnt = len(all_changes)
        sign = "+" if value >= 0 else ""
        print(f"  {GREEN}✓ Queued {cnt} SKUs {sign}{value}{kind}{RESET}")

    else:
        # Interactive arrow-key selection
        selected_sizes = select_sizes_interactive(size_map, sorted_sizes)
        if not selected_sizes:
            return (all_changes, "") if return_tag else all_changes

        # Get adjustment for all selected sizes
        print(f"\n  Enter adjustment for selected sizes ({kind}):")
        sys.stdout.write("  > ")
        sys.stdout.flush()
        adj_str = input().strip()

        try:
            value, detected_pct = parse_adj(adj_str)
            if is_pct:
                detected_pct = True
        except ValueError:
            print(f"  {RED}Invalid adjustment.{RESET}")
            return (all_changes, "") if return_tag else all_changes

        selected_set = set(selected_sizes)
        all_changes = build_changes(
            rows,
            lambda r, ss=selected_set: (r.size_num, r.unit) in ss,
            value,
            detected_pct
        )
        all_values.append(value)
        cnt = len(all_changes)
        sign = "+" if value >= 0 else ""
        sizes_str = ", ".join(f"{s[0]}{'y' if s[1]=='ปี' else 'm'}" for s in sorted(selected_sizes, key=lambda x: (x[1], x[0])))
        print(f"  {GREEN}✓ Queued {cnt} SKUs ({sizes_str}) {sign}{value}{kind}{RESET}")

    # Build tag from sizes and values
    val_sum = sum(all_values) if all_values else 0
    sign = "p" if val_sum >= 0 else "m"
    tag = f"size_{sign}{abs(val_sum):.0f}"

    return (all_changes, tag) if return_tag else all_changes


def mode_by_color(rows, is_pct, return_tag=False):
    color_map, sorted_colors = show_colors(rows)
    kind = "%" if is_pct else "฿"
    print(f"\n  Enter: {{number or color name}} {{adjustment}}")
    print(f"  e.g.:  1 +7%   or   ขาว +7%")
    print(f"  Multiple entries — one per line, blank to finish")

    all_changes = {}
    all_values = []
    while True:
        line = input("  > ").strip()
        if not line:
            break
        parts = line.rsplit(None, 1)
        if len(parts) < 2:
            print(f"  {RED}Need: color adjustment{RESET}")
            continue
        color_token = parts[0].strip()
        adj_str     = parts[1].strip()

        # resolve color
        color = None
        if color_token.isdigit():
            idx = int(color_token) - 1
            if 0 <= idx < len(sorted_colors):
                color = sorted_colors[idx]
            else:
                print(f"  {RED}Invalid number.{RESET}")
                continue
        else:
            if color_token in color_map:
                color = color_token
            else:
                print(f"  {RED}Color '{color_token}' not found.{RESET}")
                continue

        try:
            value, detected_pct = parse_adj(adj_str)
            if is_pct:
                detected_pct = True
        except ValueError:
            print(f"  {RED}Invalid adjustment.{RESET}")
            continue

        new_ch = build_changes(rows, lambda r, c=color: r.color == c, value, detected_pct)
        merge_changes(all_changes, new_ch)
        all_values.append(value)
        sign = "+" if value >= 0 else ""
        print(f"  {GREEN}✓ Queued {len(new_ch)} SKUs: {color} {sign}{value}{kind}{RESET}")

    val_sum = sum(all_values) if all_values else 0
    sign = "p" if val_sum >= 0 else "m"
    tag = f"color_{sign}{abs(val_sum):.0f}{'pct' if is_pct else 'bht'}"

    return (all_changes, tag) if return_tag else all_changes


def mode_by_sku(rows, is_pct, return_tag=False):
    kind = "%" if is_pct else "฿"
    print(f"\n  Enter SKU codes and adjustments (one per line, blank line to finish):")
    print(f"  Format: {{SKU_code}} {{adjustment}}")
    print(f"  e.g.:  3421S +7%   (partial match supported)")

    all_changes = {}
    all_values = []
    while True:
        line = input("  > ").strip()
        if not line:
            break
        parts = line.rsplit(None, 1)
        if len(parts) < 2:
            print(f"  {RED}Need: SKU adjustment{RESET}")
            continue
        sku_token = parts[0].strip()
        adj_str   = parts[1].strip()

        try:
            value, detected_pct = parse_adj(adj_str)
            if is_pct:
                detected_pct = True
        except ValueError:
            print(f"  {RED}Invalid adjustment.{RESET}")
            continue

        new_ch = build_changes(
            rows,
            lambda r, t=sku_token: t in r.sku,
            value, detected_pct
        )
        if not new_ch:
            print(f"  {YELLOW}No SKUs matched '{sku_token}'.{RESET}")
            continue
        merge_changes(all_changes, new_ch)
        all_values.append(value)
        sign = "+" if value >= 0 else ""
        print(f"  {GREEN}✓ Queued {len(new_ch)} SKUs matching '{sku_token}' {sign}{value}{kind}{RESET}")

    if all_values:
        val_sum = sum(all_values)
        tag = f"sku_{'p' if val_sum >= 0 else 'm'}{abs(val_sum):.0f}{'pct' if is_pct else 'bht'}"
    else:
        val_sum = 0
        tag = ""

    return (all_changes, tag) if return_tag else all_changes

# ── Main menu ─────────────────────────────────────────────────────────────────
SELECTION_OPTIONS = {
    "1": ("All SKUs",       ""),
    "2": ("By size",        ""),
    "3": ("By color",       ""),
    "4": ("Specific SKUs",  ""),
}

EDIT_OPTIONS = {
    "1": ("+/-% percentage (%)",  True),
    "2": ("+/- flat amount (฿)",  False),
}

GROUP_MAP = {
    "1": "all",
    "2": "size",
    "3": "color",
    "4": "sku",
}

def select_selection_arrow():
    """Arrow-key menu: selection type."""
    options = [
        ("1", "All SKUs",       ""),
        ("2", "By size",        ""),
        ("3", "By color",       ""),
        ("4", "Specific SKUs",  ""),
    ]
    return select_menu_arrow(options, "SELECT SKUs (↑↓ navigate, Enter select, Esc cancel)")

def select_edit_type_arrow():
    """Arrow-key menu: edit type."""
    options = [
        ("1", "+/- percentage (%)",  ""),
        ("2", "+/- flat amount (฿)", ""),
    ]
    return select_menu_arrow(options, "SELECT EDIT TYPE (↑↓ navigate, Enter select, Esc cancel)")

def run_mode(selection_key, edit_key, rows):
    is_pct = EDIT_OPTIONS[edit_key][1]
    grp = GROUP_MAP[selection_key]
    label = SELECTION_OPTIONS[selection_key][0]
    edit_type = "pct" if is_pct else "flat"
    print(f"\n  {BOLD}{label} → {'percentage (%)' if is_pct else 'flat amount (฿)'}{RESET}")
    if grp == "all":
        changes, val = mode_all(rows, is_pct, return_val=True)
        sign = "p" if val >= 0 else "m"
        tag = f"{grp}_{sign}{abs(val):.0f}{edit_type}"
        return changes, tag
    elif grp == "size":
        changes, tag = mode_by_size(rows, is_pct, return_tag=True)
        return changes, tag
    elif grp == "color":
        changes, tag = mode_by_color(rows, is_pct, return_tag=True)
        return changes, tag
    elif grp == "sku":
        changes, tag = mode_by_sku(rows, is_pct, return_tag=True)
        return changes, tag
    return {}, ""

# ── Startup banner ────────────────────────────────────────────────────────────
def print_banner():
    print(f"\n{CYAN}╔══════════════════════════════════════════╗{RESET}")
    print(f"{CYAN}║{RESET}  {BOLD}Shopee Mass Price Editor{RESET}                {CYAN}║{RESET}")
    print(f"{CYAN}║{RESET}  Build: {BOLD}{BUILD_LABEL}{RESET}           {CYAN}║{RESET}")
    print(f"{CYAN}╚══════════════════════════════════════════╝{RESET}")

def print_file_info(path, rows):
    n_products = unique_products(rows)
    n_skus     = sum(1 for r in rows if r.price is not None)
    mn, mx     = price_range(rows)
    fname      = os.path.basename(path)
    print(f"\n  {GREEN}✓ Loaded: {fname}{RESET}")
    print(f"    Products:    {n_products}")
    print(f"    SKUs:        {n_skus}")
    if mn is not None:
        print(f"    Price range: ฿{mn:.0f} – ฿{mx:.0f}")

# ── Post-save legend ──────────────────────────────────────────────────────────
def print_legend():
    print(f"\n  {BOLD}Color legend:{RESET}")
    def swatch(hex_c):
        return f"\033[48;2;{int(hex_c[0:2],16)};{int(hex_c[2:4],16)};{int(hex_c[4:6],16)}m  \033[0m"
    print(f"  {swatch('1B5E20')} Dark green   = large increase   (>20%)")
    print(f"  {swatch('2E7D32')} Medium green = moderate increase (10–20%)")
    print(f"  {swatch('66BB6A')} Mid green    = increase          (5–10%)")
    print(f"  {swatch('A5D6A7')} Light green  = small increase    (2–5%)")
    print(f"  {swatch('E8F5E9')} Pale green   = tiny increase     (<2%)")
    print(f"  {swatch('7F0000')} Dark red     = large decrease   (>20%)")
    print(f"  {swatch('C62828')} Medium red   = moderate decrease (10–20%)")
    print(f"  {swatch('EF5350')} Mid red      = decrease          (5–10%)")
    print(f"  {swatch('EF9A9A')} Light red    = small decrease    (2–5%)")
    print(f"  {swatch('FFEBEE')} Pale red     = tiny decrease     (<2%)")
    print(f"  {swatch('FFE0B2')} Orange       = price capped by Shopee limits")

# ── Main editor loop ──────────────────────────────────────────────────────────
def editor_loop(file_path):
    try:
        wb, ws, rows = load_file(file_path)
    except FileNotFoundError:
        print(f"  {RED}File not found: {file_path}{RESET}")
        return False
    except Exception as e:
        print(f"  {RED}Error loading file: {e}{RESET}")
        return False

    data_rows = [r for r in rows if any(
        [r.product_id, r.product_name, r.sku])]
    if not data_rows:
        print(f"  {RED}No product rows found in this file.{RESET}")
        return False
    row_by_idx = {r.row_idx: r for r in data_rows}

    print_file_info(file_path, data_rows)

    # Track original prices (first load) and cumulative current prices
    original_prices = {r.row_idx: r.price for r in data_rows if r.price is not None}
    current_prices = {r.row_idx: r.price for r in data_rows if r.price is not None}
    all_changes = {}  # row_idx -> (original, current, capped)

    while True:
        sel_key = select_selection_arrow()
        if sel_key is None:
            print(f"  {DIM}Cancelled.{RESET}")
            continue
        if sel_key == "q":
            print(f"  {DIM}Goodbye.{RESET}")
            return False

        edit_key = select_edit_type_arrow()
        if edit_key is None:
            print(f"  {DIM}Cancelled.{RESET}")
            continue
        if edit_key == "q":
            continue

        changes, _ = run_mode(sel_key, edit_key, data_rows)
        if not changes:
            print(f"  {YELLOW}No changes computed.{RESET}")
            continue

        changed_count = sum(1 for idx, (o, n, _) in changes.items() if o != n)
        if changed_count == 0:
            print(f"  {YELLOW}All prices unchanged (same values).{RESET}")
            continue

        show_preview(data_rows, changes)

        ans = input(f"\n  Apply? [y/N]: ").strip().lower()
        if ans != "y":
            print(f"  {DIM}Discarded. Returning to menu.{RESET}")
            continue

        # Merge changes into cumulative tracking
        for idx, (_, new_p, capped) in changes.items():
            orig = original_prices.get(idx)
            if orig is None:
                continue
            # Update current prices
            current_prices[idx] = new_p
            all_changes[idx] = (orig, new_p, capped)
            # Also apply to workbook
            row = row_by_idx.get(idx)
            if row is not None:
                row.price = new_p

        print(f"\n  {GREEN}✓ Applied changes. {len(all_changes)} total SKUs modified so far.{RESET}")

        again = input(f"\n  Run another adjustment? [y/N]: ").strip().lower()
        if again == "y":
            continue
        else:
            # Save final file and summary report
            if not all_changes:
                print(f"  {DIM}No changes to save.{RESET}")
                return True

            # Apply all changes to workbook and save
            final_changes = {idx: (orig, curr, capped) for idx, (orig, curr, capped) in all_changes.items()}
            out_path = make_output_path(file_path)
            try:
                save_file(wb, ws, data_rows, final_changes, out_path)
            except Exception as e:
                print(f"  {RED}Save failed: {e}{RESET}")
                return False

            fname = os.path.basename(out_path)
            print(f"\n  {GREEN}✓ Saved: {fname}{RESET}")
            print_legend()

            # Save summary report
            try:
                summary_path = save_summary_report(original_prices, current_prices, data_rows, out_path)
                print(f"  {GREEN}✓ Summary report: {os.path.basename(summary_path)}{RESET}")
            except Exception as e:
                print(f"  {YELLOW}Summary report failed: {e}{RESET}")

            print(f"  {DIM}Done.{RESET}")
            return True

# ── Self-tests ────────────────────────────────────────────────────────────────
def run_tests():
    print(f"\n{BOLD}{'═'*50}{RESET}")
    print(f"  {BOLD}Running self-tests…{RESET}")
    print(f"{'═'*50}")
    failures = 0

    def assert_eq(desc, got, expected):
        nonlocal failures
        ok = abs(got - expected) < 0.001 if isinstance(got, float) else got == expected
        status = f"{GREEN}PASS{RESET}" if ok else f"{RED}FAIL{RESET}"
        print(f"  [{status}] {desc}: got={got!r}  expected={expected!r}")
        if not ok:
            failures += 1

    # price calc pct
    p, c = calc_price_pct(289, 7)
    assert_eq("pct +7 on 289", p, 309.23)
    p, c = calc_price_pct(289, -5)
    assert_eq("pct -5 on 289", p, 274.55)
    p, c = calc_price_pct(1, -50)
    assert_eq("pct clamped to min", p, 1.0)
    _, capped = calc_price_pct(490000, 10)
    assert_eq("pct clamped to max — capped flag", capped, True)

    # flat calc
    p, c = calc_price_flat(289, 20)
    assert_eq("flat +20 on 289", p, 309.0)
    p, c = calc_price_flat(289, -15)
    assert_eq("flat -15 on 289", p, 274.0)

    # variation parsing
    r = parse_variation("น้ำเงิน 3469S-45,6ปี 20-22kg")
    assert_eq("parse color น้ำเงิน", r[0], "น้ำเงิน")
    assert_eq("parse size 6", r[1], 6)
    assert_eq("parse unit ปี", r[2], "ปี")

    r = parse_variation("ขาว 3469S-18,24เดือน 12-13kg")
    assert_eq("parse color ขาว", r[0], "ขาว")
    assert_eq("parse size 24", r[1], 24)
    assert_eq("parse unit เดือน", r[2], "เดือน")

    r = parse_variation("เขียวละมุน 3421S-26,14ปี 34-36kg")
    assert_eq("parse color เขียวละมุน", r[0], "เขียวละมุน")
    assert_eq("parse size 14", r[1], 14)

    r = parse_variation("bad input no comma")
    assert_eq("parse failure returns None", r, None)

    # Format 1: SKU-first
    r = parse_variation("2669S-21 แดง,3ปี 14-15kg")
    assert_eq("Format1 color แดง", r[0], "แดง")
    assert_eq("Format1 size 3", r[1], 3)
    assert_eq("Format1 unit ปี", r[2], "ปี")

    # Format 3: no SKU (plain color)
    r = parse_variation("เขียว,3ปี 14-15kg")
    assert_eq("Format3 color เขียว", r[0], "เขียว")
    assert_eq("Format3 size 3", r[1], 3)
    assert_eq("Format3 unit ปี", r[2], "ปี")

    # Format 3: no SKU (product type name)
    r = parse_variation("ชุดเดรสเด็กผู้หญิง,12เดือน 8-9kg")
    assert_eq("Format3 product-type color", r[0], "ชุดเดรสเด็กผู้หญิง")
    assert_eq("Format3 product-type size 12", r[1], 12)
    assert_eq("Format3 product-type unit เดือน", r[2], "เดือน")

    # Format 4: pure-number code + Y suffix
    r = parse_variation("สีฟ้า 45,8Y")
    assert_eq("Format4 color สีฟ้า", r[0], "สีฟ้า")
    assert_eq("Format4 size 8", r[1], 8)
    assert_eq("Format4 unit ปี (Y suffix)", r[2], "ปี")

    # parse_adj
    v, p = parse_adj("+7%")
    assert_eq("parse_adj +7%", v, 7.0)
    assert_eq("parse_adj pct True", p, True)
    v, p = parse_adj("-15")
    assert_eq("parse_adj -15", v, -15.0)
    assert_eq("parse_adj pct False", p, False)

    # fill bands
    f, _ = get_fill(100, 125)    # 25% increase → dark green
    assert_eq("fill 25% increase fgColor", f.fgColor.rgb[-6:].upper(), "1B5E20")
    f, _ = get_fill(100, 70)     # 30% decrease → #7F0000
    assert_eq("fill 30% decrease fgColor", f.fgColor.rgb[-6:].upper(), "7F0000")

    print(f"\n  {'─'*46}")
    if failures == 0:
        print(f"  {GREEN}{BOLD}All tests passed.{RESET}")
    else:
        print(f"  {RED}{BOLD}{failures} test(s) failed.{RESET}")
    print(f"{'═'*50}\n")
    return failures == 0

# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    if "--test" in sys.argv:
        ok = run_tests()
        sys.exit(0 if ok else 1)

    print_banner()

    while True:
        file_path = input("\n  Enter path to .xlsx file: ").strip().strip('"').strip("'")
        if not file_path:
            continue
        if not os.path.exists(file_path):
            print(f"  {RED}File not found: {file_path}{RESET}")
            continue
        if not file_path.lower().endswith(".xlsx"):
            print(f"  {YELLOW}Warning: file does not end in .xlsx{RESET}")
        break

    editor_loop(file_path)
    print(f"\n  {DIM}python price_editor.py{RESET}\n")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n  {RED}ERROR: {e}{RESET}")
        import traceback; traceback.print_exc()
        print(f"\n  {YELLOW}Press Enter to exit...{RESET}")
        try: input()
        except: pass
